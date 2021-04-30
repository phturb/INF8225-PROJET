import matplotlib.pyplot as plt
import openpyxl
from pathlib import Path

def smooth(scalars: [], weight: float):  # Weight between 0 and 1
    last = scalars[0]  # First value in the plot (first timestep)
    smoothed = []
    for point in scalars:
        smoothed_val = last * weight + (1 - weight) * point  # Calculate smoothed value
        smoothed.append(smoothed_val)                        # Save it
        last = smoothed_val                                  # Anchor the last smoothed value

    return smoothed


#################### DATA #####################
data = {}

xlsx_file = Path('merged.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file) 

agent = []
names = []
color = ['#0000ff','#00ff00','#ff0000','#ff9100','#fbff00','#6fff00','#00ffc8', '#c800ff', '#ff00dd', '#00ddff']

for column in wb_obj.active.iter_cols(1, wb_obj.active.max_column):
    names.append(column[0].value)

for name in wb_obj.sheetnames:
    data[name] = {}
    agent.append(name)

for n, sheet in enumerate(wb_obj):
    data[agent[n]]['color'] = color[n]
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            data[agent[n]][row[0]] = []
            data[agent[n]][row[1]] = []
            data[agent[n]][row[2]] = []

        else:
            data[agent[n]][names[0]].append(row[0])
            data[agent[n]][names[1]].append(row[1])
            data[agent[n]][names[2]].append(row[2])

#################### SMOOTHING #####################
weight = 0.75
for ag in agent:
    data[ag][names[2]] = smooth(data[ag][names[2]], weight)

#################### SIMPLE #####################
for ag in agent:
    curAg = data[ag]

    fig, ax = plt.subplots(figsize=(11, 8))
    for label in (ax.get_xticklabels() + ax.get_yticklabels()):
        label.set_fontsize(16)

    plt.xlabel("Pas", fontsize=18)
    plt.ylabel("Récompenses", fontsize=16)
    plt.ylim(0, 10)

    lb = ag
    if lb == "rainbow no noisy":
        lb = "rainbow\nno noisy"
        
    ax.plot(curAg[names[1]], curAg[names[2]], color=curAg['color'], label=lb)

    leg = ax.legend(fontsize=12, loc="upper left")
    # plt.show()
    plt.savefig(f"{ag}.png")

#################### AVEC DQN #####################
for ag in agent:
    if ag != 'dqn':
        dqn = data['dqn']
        curAg = data[ag]

        fig, ax = plt.subplots(figsize=(11, 8))
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(16)

        plt.xlabel("Pas", fontsize=18)
        plt.ylabel("Récompenses", fontsize=16)
        plt.ylim(0, 10)

        lb = ag
        if lb == "rainbow no noisy":
            lb = "rainbow\nno noisy"
            
        ax.plot(dqn[names[1]], dqn[names[2]], color=dqn['color'], label=lb)
        ax.plot(curAg[names[1]], curAg[names[2]], color=curAg['color'], label=lb)

        leg = ax.legend(fontsize=12, loc="upper left")
        # plt.show()
        plt.savefig(f"{ag}vsdqn.png")


#################### TOUT #####################
fig, ax = plt.subplots(figsize=(11, 8))
for label in (ax.get_xticklabels() + ax.get_yticklabels()):
    label.set_fontsize(16)

plt.xlabel("Pas", fontsize=18)
plt.ylabel("Récompenses", fontsize=16)
plt.ylim(0, 210)
for ag in agent:
    curAg = data[ag]
    lb = ag
    if lb == "rainbow no noisy":
        lb = "rainbow\nno noisy"
    ax.plot(curAg[names[1]], curAg[names[2]], color=curAg['color'], label=lb)

leg = ax.legend(fontsize=12, loc="upper left")
# plt.show()
plt.savefig("all.png")